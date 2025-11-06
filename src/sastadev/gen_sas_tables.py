from collections import defaultdict
from openpyxl import load_workbook
import os
from sastadev.conf import settings
from sastadev.constants import actualfolder, figuresfolder, overallsasfolder
from sastadev.datasets import trainingdatasets, testdatasets, alldatasets
from sastadev.sastatypes import HeadedTable, MethodName, Table
from sastadev.xlsx import getxlsxdata, mkworkbook
from typing import Callable, List, Tuple

eps = ''
Deltasym = '\u0394'

figures_dataset_col = 0
figures_method_col = 1
figures_avg_f1_col = 4
figures_min_f1_col = 5
figures_max_f1_col = 6
figures_sf1_col = 7

overall_sas_dataset_col = 0
overall_sas_condition_col = 1
overall_sas_status_col = 2
overall_sas_ok_col = 3
overall_sas_orig_avf_f1_col = 4

shortsummary_samplecount_col = 1
mindelta_col = 3

summary_ok_col = 3
summary_orig_avg_f1_col = 4
summary_min_utt_col  = 10

trainingdatasetnames = [ds.name for ds in trainingdatasets]
testdatasetnames = [ds.name for ds in testdatasets]
excluded_datasets = ['eliska', 'schlichtingall', 'vklasta_ang']

sasta_performance_header = ['Dataset', 'Use', 'Method', 'Avg F1', 'Min F1', 'Max F1']

outpath = r'D:\Dropbox\jodijk\presentations\2026\LREC\tables'
sasta_figures_fn = 'SASTA_Figures.xlsx'
sas_ref_header  = ['Use', 'Dataset', 'Method', '#', 'OK', 'Orig Avg F1',
                   'Avg F1', 'Min F1', 'Max F1',
                   'Avg #Utt', 'Min #Utt', 'Max #Utt']
deltaheader = [f'Min F1 {Deltasym}', f'Max F1 {Deltasym}', f'Avg F1 {Deltasym}']

cond_count_header = ['Use', 'Dataset', 'Method', '#', 'OK']

trainortestdatasets = trainingdatasets + testdatasets

# this must move to xlsx
def read_ws(ws, headerrow=0) -> HeadedTable:
    header = []
    data = []
    rowctr = -1
    for row in ws.iter_rows():
        rowctr += 1
        #preheader
        if rowctr < headerrow:
            continue
        # header
        elif rowctr == headerrow:
            header = [eps if cell.value is None else cell.value for cell in row]
        else:
            valuerow = [eps if cell.value is None else cell.value for cell in row]
            data.append(valuerow)
    return header, data

def get_sasta_cond_count_dict(datefolder: str, criterion: Callable) -> dict:
    resultdict = defaultdict(int)
    infilename = sasta_figures_fn
    inpath = os.path.join(settings.DATAROOT, figuresfolder, datefolder)
    infullname = os.path.join(inpath, infilename)
    try:
        wb = load_workbook(infullname)
    except Exception:
        return resultdict
    wsns = wb.sheetnames[1:]
    for wsn in wsns:
        ws = wb[wsn]
        header, data = read_ws(ws)
        samplecount = 0
        okcount = 0
        for row in data:
            rawsf1 = row[figures_sf1_col]
            sf1 = eval(rawsf1)
            if not isinstance(sf1, float):
                print(f'Unexpected value for sf1 ({sf1}) in {wsn}: {str(row)} ')
                continue
            samplecount += 1
            if criterion(sf1):
                okcount += 1
        resultdict[wsn] = okcount, samplecount
    return resultdict

def get_sasta_cond_headed_table(datefolder: str, criterion: Callable) -> HeadedTable:
    data = []
    cond_count_dict = get_sasta_cond_count_dict(datefolder, criterion)
    for dsn in cond_count_dict:
        if dsn in excluded_datasets:
            continue
        okcount, samplecount = cond_count_dict[dsn]
        if dsn in trainingdatasetnames:
            use = 'train'
        elif dsn in testdatasetnames:
            use = 'test'
        methodname = get_method_name_of_dsname(dsn)
        row = [use, dsn, methodname, samplecount, okcount]
        data.append(row)
    return cond_count_header, data





def get_sasta_performance_table_f1(datefolder: str) -> HeadedTable:
    # Dataset Use Method Avg F1 Min F1 Max F1
    infilename = sasta_figures_fn
    inpath = os.path.join(settings.DATAROOT, figuresfolder, datefolder)
    infullname = os.path.join(inpath, infilename)
    header, data = getxlsxdata(infullname)
    trainingrows = []
    testrows = []
    for row in data:
        datasetname = row[figures_dataset_col]
        methodname = row[figures_method_col]
        avg_f1 = row[figures_avg_f1_col]
        min_f1 = row[figures_min_f1_col]
        max_f1 = row[figures_max_f1_col]
        if datasetname in excluded_datasets:
            continue
        if datasetname in trainingdatasetnames:
            use = 'train'
        elif datasetname in testdatasetnames:
            use = 'test'
        else: use = ''
        if use in ['train', 'test']:
            newrow = [datasetname, use, methodname, round(avg_f1,1), round(min_f1,1), round(max_f1, 1)]
            if use == 'train':
                trainingrows.append(newrow)
            elif use == 'test':
                testrows.append(newrow)
    table = trainingrows + testrows
    return sasta_performance_header, table


def get_shortsummary_dict(shortsummarydata: Table) -> dict:
    resultdict = {row[0]: row for row in shortsummarydata}
    return resultdict

def get_method_name_of_dsname(dsname: str) -> MethodName:
    for ds in trainortestdatasets:
        if ds.name == dsname and ds.name not in excluded_datasets:
            return ds.method

# add options for separate tables; add option for including delta columns
def get_sas_data(datefolder=None, reference=True, include_delta=False, combined=True) -> Tuple[List, Table, Table]:
    trainfilename = 'Overall_SAS.xlsx'
    testfilename = 'Overall_SAS_testdata.xlsx'
    basepath = os.path.join(settings.DATAROOT, overallsasfolder)
    if datefolder is None:
        inpath = basepath
    else:
        inpath = os.path.join(basepath, datefolder)
    trainfullname = os.path.join(inpath, trainfilename)
    testfullname = os.path.join(inpath, testfilename)
    trainrows = get_sas_rows(trainfullname, use='train', reference=reference, include_delta=include_delta)
    testrows = get_sas_rows(testfullname, use='test', reference=reference, include_delta=include_delta)
    if combined:
        rows1 = trainrows + testrows
        rows2 = []
        header = sas_ref_header
    else:
        # if sepaarte we do not need the column use (colomn 0)
        rows1 = [row[1:] for row in trainrows]
        rows2 = [row[1:] for row in testrows]
        header = sas_ref_header[1:]
    if include_delta:
        header += deltaheader
    else:
        pass
    return header, rows1, rows2

def get_sas_rows(fullname: str, use='train', reference=True, include_delta=False) -> Table:
    if reference:
        status = 'Target'
    else:
        status = 'Results'
    header, data = getxlsxdata(fullname, sheetname='Summary')
    shortsummaryheader, shortsummarydata = getxlsxdata(fullname, sheetname='ShortSummary')
    shortsummary_dict = get_shortsummary_dict(shortsummarydata)
    resultrows =[]
    for row in data:
        datasetname = row[overall_sas_dataset_col]
        if datasetname not in excluded_datasets and \
           row[overall_sas_condition_col] == 'Restricted' and \
           row[overall_sas_status_col] == status:
            samplecount = shortsummary_dict[datasetname][shortsummary_samplecount_col]
            methodname = get_method_name_of_dsname(datasetname)
            resultsrow = [row[summary_ok_col]] + \
                         [round(cell,1) for cell in row[summary_orig_avg_f1_col:summary_min_utt_col]] + \
                         row[summary_min_utt_col:]
            newrow = [use, datasetname, methodname, samplecount] + resultsrow
            if include_delta:
                rawdeltarow = shortsummary_dict[datasetname][mindelta_col:]
                deltarow = [round(cell,1) for cell in rawdeltarow]
                newrow += deltarow
            resultrows.append(newrow)
    return resultrows

if __name__ == '__main__':
    header, data = get_sasta_performance_table_f1(datefolder='actual')
    outfullname = os.path.join(outpath, sasta_figures_fn)
    wb = mkworkbook(outfullname, [header], data, freeze_panes=(1,0))
    wb.close()

    header, data, _ = get_sas_data(datefolder=None, reference=True, combined=True)
    sasref_filename = 'SAS Reference.xlsx'
    outfullname = os.path.join(outpath, sasref_filename)
    wb = mkworkbook(outfullname, [header], data, freeze_panes=(1,0))
    wb.close()

    header, traindata, testdata = get_sas_data(datefolder=None, reference=False, include_delta=True, combined=False)
    sastrain_filename = 'SAS_train_results.xlsx'
    outfullname = os.path.join(outpath, sastrain_filename)
    wb = mkworkbook(outfullname, [header], traindata, freeze_panes=(1,0))
    wb.close()

    sastest_filename = 'SAS_test_results.xlsx'
    outfullname = os.path.join(outpath, sastest_filename)
    wb = mkworkbook(outfullname, [header], testdata, freeze_panes=(1,0))
    wb.close()

    criterion = lambda x: x >= 95.0
    header, data = get_sasta_cond_headed_table(datefolder='actual', criterion=criterion)
    cond_count_filename = 'Samples_above.xlsx'
    outfullname = os.path.join(outpath, cond_count_filename)
    wb = mkworkbook(outfullname, [header], data, freeze_panes=(1,0))
    wb.close()
