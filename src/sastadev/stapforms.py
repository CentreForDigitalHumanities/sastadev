import os
from collections import defaultdict
from io import BytesIO
from shutil import copyfile, copyfileobj

from openpyxl import load_workbook

from sastadev.allresults import AllResults, mkresultskey
from sastadev.conf import settings
from sastadev.forms import getformfilename

scoresheetname = 'STAP 1 - 5'
extrasheetname = 'extra'
maxutt = 50
zerocount = 0
basexl = os.path.join(settings.SD_DIR, 'data', 'form_templates', 'STAP Excel VUmc 2018.xlsx')

NS = mkresultskey('S001')
OS = mkresultskey('S002')
PV = mkresultskey('S003')
SGG = mkresultskey('S004')
VT = mkresultskey('S005')
VD = mkresultskey('S006')
N = mkresultskey('S007')
BvBep = mkresultskey('S008')
zelfvnw3 = mkresultskey('S009')
BBp = mkresultskey('S010')
BBt = mkresultskey('S011')
BBo = mkresultskey('S012')

AG = 33
Ucol = 21
AF = 32

# order in the Excel sheet: NS	OS	PV	SGG	VT	VD	N	BvBep	zelf. vnw. 3	BB p	BB t	BB o
# i.e.
sorteditemlist = [NS, OS, PV, SGG, VT, VD, N, BvBep, zelfvnw3, BBp, BBt, BBo]


def data2rowtuples(data):
    # data is  a dictionary with key item and as value a counter with (uttid, count) items
    newdata = defaultdict(lambda: defaultdict(int))
    for item in data:
        for (uttid, count) in data[item].items():
            newdata[uttid][item] += count

    rowlist = []
    uttidlist = [uttid for uttid in newdata]
    sorteduttidlist = sorted(uttidlist)

    for uttid in sorteduttidlist:
        row = []
        for item in sorteditemlist:
            if item in newdata[uttid]:
                row.append(newdata[uttid][item])
            else:
                row.append(zerocount)
        rowlist.append((uttid, row))

    return rowlist


def makestapform(allresults, _, basexl=basexl, in_memory=False):
    if not in_memory:
        # copy the basexl to a new one with the appropriate name
        #(base, ext) = os.path.splitext(allresults.filename)
        #target = base + '_STAP-Form' + '.xlsx'
        target = getformfilename(allresults.filename, '_STAP-Form')

        copyfile(basexl, target)

        # open the workbook
        wb = load_workbook(filename=target)
    else:
        target = BytesIO()
        with open(basexl, 'rb') as source:
            copyfileobj(fsrc=source, fdst=target)
        wb = load_workbook(target)

    # gather the results

    # put the results in the right order
    rowlist = data2rowtuples(allresults.coreresults)

    ws = wb[scoresheetname]

    cols = ['U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']
    # adapt the relevant sheet
    for (uttid, row) in rowlist:
        uttidrow = int(uttid) + 3
        xluttctr = ws.cell(column=AG, row=uttidrow).value
        uttidrowstr = str(uttidrow)
        if int(uttid) == xluttctr:
            for col, el in zip(cols, row):
                # special proviso for PV in column W
                if col == 'W':
                    el = el - 1
                cellkey = col + uttidrowstr
                ws[cellkey] = el
        else:
            settings.LOGGER.error('Unexpected utterance id encountered: {}'.format(uttid))

    # add the communicative and noncommunicative word counts to the form, restrict it to max 50
    found_utt_count = len(allresults.commwordcounts)
    if  found_utt_count != 50:
        settings.LOGGER.warning(f'{found_utt_count} utterances encountered. Results unreliable')
    if found_utt_count > 50:
        settings.LOGGER.warning(f'Utterances limited to 50. Results unreliable')

    uttidcol = 'A'
    countcol = 'C'
    startrow = 4
    currentrow = startrow
    for (uttid, count) in allresults.commwordcounts[:50]:
        uttidcellkey = uttidcol + str(currentrow)
        countcellkey = countcol + str(currentrow)
        ws[uttidcellkey] = uttid
        ws[countcellkey] = count
        currentrow += 1

    countcol = 'B'
    startrow = 4
    currentrow = startrow
    for (uttid, count) in allresults.noncommwordcounts[:50]:
        countcellkey = countcol + str(currentrow)
        ws[countcellkey] = count
        currentrow += 1

    # find the 5 longest utterances
    sorted_commwordcounts = sorted(allresults.commwordcounts, key=lambda x: x[1], reverse=True)
    top5 = sorted_commwordcounts[:5]

    wsextra = wb[extrasheetname]
    colids = ['D', 'E', 'F', 'G', 'H']
    for i, el in enumerate(top5):
        col = colids[i]
        cellkey = f'{col}11'
        wsextra[cellkey] = el[1]



    # save the workbook
    wb.save(target)
    wb.close()

    # return the workbook- not needed
    return target


def test():
    coreresults = {NS: {'1': 3}, OS: {'1': 2, '2': 6}}
    postresults = {}
    allmatches = {}
    fn = 'STAP42.xml'
    analysedtrees = {}
    allresults = AllResults(0, coreresults, postresults, allmatches, fn, analysedtrees)
    fnbase, _ = os.path.splitext(fn)
    formxl = fnbase + '_form' + '.xlsx'
    makestapform(allresults, _)


if __name__ == '__main__':
    test()
